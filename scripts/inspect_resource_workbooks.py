from pathlib import Path
import json

from openpyxl import load_workbook


PLAN = Path(r"D:/pm/resource/1/计划工时按工作-STK.xlsx")
ACTUAL = Path(r"D:/pm/resource/1/填报工时原始数据_STK.xlsx")


def summarize(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    info = {"file": str(path), "sheets": wb.sheetnames, "samples": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
            rows.append(list(row[:20]))
        info["samples"][name] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "rows": rows,
        }
    wb.close()
    return info


print(json.dumps({"plan": summarize(PLAN), "actual": summarize(ACTUAL)}, ensure_ascii=False, indent=2, default=str))
